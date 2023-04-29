from openpyxl import *
from random import *

libro_excel = Workbook()

hoja_excel = libro_excel.active

try:
    hoja_excel['A1'] = "product_id"
    hoja_excel['B1'] = "product_name"
    hoja_excel['C1'] = "qty"
    hoja_excel['D1'] = "price"

    productos = [
    {'product_id': 'p01', 'product_name': 'arroz', 'qty': 2, 'price': 3.5},
    {'product_id': 'p01', 'product_name': 'arroz', 'qty': 1, 'price': 3.5},
    {'product_id': 'p01', 'product_name': 'arroz', 'qty': 3, 'price': 3.5},
    {'product_id': 'p02', 'product_name': 'azucar', 'qty': 4, 'price': 4.1},
    {'product_id': 'p03', 'product_name': 'aceite', 'qty': 2, 'price': 8.6},
    {'product_id': 'p04', 'product_name': 'leche', 'qty': 1, 'price': 3.7}]

    for i,producto in enumerate (productos, start=2):
        hoja_excel.cell(row = i, column = 1, value=producto['product_id'])
        hoja_excel.cell(row = i, column = 2, value=producto['product_name'])
        hoja_excel.cell(row = i, column = 3, value=producto['qty'])
        hoja_excel.cell(row = i, column = 4, value=producto['price'])
        
except:
    pass

libro_excel.save('nuevo_libro.xlsx')
