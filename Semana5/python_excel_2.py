import openpyxl
from openpyxl import *

try:
    libro_excel = openpyxl.load_workbook('Ejemplo2.xlsx')
    hoja_excel = libro_excel.active

    etiqueta_1 = hoja_excel['A1'].value
    etiqueta_2 = hoja_excel['A2'].value
    etiqueta_3 = hoja_excel['A3'].value
    etiqueta_4 = hoja_excel['A4'].value


    libro_excel2 = Workbook()
    hoja_excel2 = libro_excel2.active
    hoja_excel2['A1'] = etiqueta_1
    hoja_excel2['B1'] = etiqueta_2
    hoja_excel2['C1'] = etiqueta_3
    hoja_excel2['D1'] = etiqueta_4

    for i in range(2, 6):
        hoja_excel2.cell(row=i,column=1).value = "LG"
        hoja_excel2.cell(row=i, column=2).value = "RPP"
        hoja_excel2.cell(row=i, column=3).value = "Samsung"
        hoja_excel2.cell(row=i, column=4).value = "Samsung"


    libro_excel2.save("ejemplo03.xlsx")

except:
    pass


