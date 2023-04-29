import openpyxl

ejemplo3 = openpyxl.load_workbook('ejemplo03.xlsx')

hoja = ejemplo3.active

etiqueta_1 = hoja['A1'].value
etiqueta_2 = hoja['A2'].value
etiqueta_3 = hoja['A3'].value
etiqueta_4 = hoja['A4'].value

for i in range(2, 6):
    tv = hoja.cell(row=i, column=1).value
    radio = hoja.cell(row=i, column=2).value
    refrigeradora = hoja.cell(row=i, column=3).value
    lavadora = hoja.cell(row=i, column=4).value
    print(f"{tv} | {radio} | {refrigeradora} | {lavadora}")
